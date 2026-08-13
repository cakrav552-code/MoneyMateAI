import os
import sqlite3
from datetime import datetime

from telegram import Update
from telegram.ext import ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = sqlite3.connect("data.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, jenis, kategori, keterangan, nominal, tanggal
        FROM transaksi
        ORDER BY id ASC
    """)

    data = cursor.fetchall()
    conn.close()

    if not data:
        await update.message.reply_text(
            "📊 Belum ada transaksi untuk diekspor."
        )
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"

    # Header
    headers = [
        "ID",
        "Jenis",
        "Kategori",
        "Keterangan",
        "Nominal",
        "Tanggal"
    ]

    ws.append(headers)

    # Data
    for row in data:
        ws.append(row)

    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )
        cell.alignment = Alignment(horizontal="center")

    # Format nominal
    for cell in ws["E"][1:]:
        cell.number_format = '"Rp" #,##0'

    # Filter
    ws.auto_filter.ref = ws.dimensions

    # Freeze header
    ws.freeze_panes = "A2"

    # Lebar kolom otomatis
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column_letter].width = min(
            max_length + 2,
            40
        )

    nama_file = (
        f"MoneyMate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb.save(nama_file)

    try:
        with open(nama_file, "rb") as file:
            await update.message.reply_document(
                document=file,
                filename=nama_file,
                caption="📊 Export transaksi berhasil."
            )
    finally:
        if os.path.exists(nama_file):
            os.remove(nama_file)
